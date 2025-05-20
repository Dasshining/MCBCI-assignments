import openpyxl
import random

# Cargar el archivo y la hoja
wb = openpyxl.load_workbook("golf-dataset-categorical.xlsx")
ws = wb.active

# Obtener los encabezados (primera fila)
headers = [cell.value for cell in ws[1]]

# Crear lista de diccionarios para cada fila de datos
dic_list = []
for row in ws.iter_rows(min_row=2, values_only=True):
    dic = dict(zip(headers, row))
    dic_list.append(dic)

# Imprimir la lista de diccionarios
def print_dic_list(dic_list):
    for dic in dic_list:
        print(dic)
    print("\n")

def split_data(data, probability=50):
    training_lines = int(len(data) * (probability / 100))
    test_lines = int(len(data) - training_lines)
    # Shuffle the data
    random.shuffle(data)
    # Split the data into training and test sets
    training_dict = data[:training_lines]
    test_dict = data[training_lines:training_lines + test_lines]
    return training_dict, test_dict

def build_frequency_table(data, target):
    features = [key for key in data[0] if key != target]
    target_values = set(row[target] for row in data)

    freq_table = {}
    for feature in features:
        freq_table[feature] = {}
        feature_values = set(row[feature] for row in data)
        for value in feature_values:
            freq_table[feature][value] = {t_val: 0 for t_val in target_values}

    for row in data:
        for feature in features:
            value = row[feature]
            t_val = row[target]
            freq_table[feature][value][t_val] += 1
    return freq_table

def train_one_r(freq_table):
    # Train the OneR model using the frequency table

    # Initialize variables to track the best feature and its error
    best_feature = None
    best_probability = None
    best_rule = {}

    # Iterate over each feature in the frequency table
    for feature, values in freq_table.items():
        print(f"Feature: {feature}")
        sum_target_value = 0
        sum_target_per_feature = 0
        # Iterate over each value of the feature
        for value, counts in values.items():
            temp_target_value = None
            print(f"    Value: {value}, Counts: {counts}")
            # Iterate over each target value and its count
            for target_value, count in counts.items():
                # Retrieve the total count of the target values
                sum_target_value += count
                # Check if the current count is less than the temporary target value
                if temp_target_value == None:
                    temp_target_value = count
                elif count < temp_target_value:
                    temp_target_value = count
                # print(f"        Target value: {target_value}, Count: {count}")
                # print(f"        temp: {temp_target_value}")
            # Calculate the sum of target values for the feature    
            sum_target_per_feature += temp_target_value
        print(f"{feature} probability: {sum_target_per_feature}/{sum_target_value}")
        # Check if this is the first feature being evaluated
        if best_feature == None:
            best_feature = feature
            best_probability = sum_target_per_feature
        else:
            # Check if the current feature has a lower error than the best feature so far
            if sum_target_per_feature < best_probability:
                best_feature = feature
                # Update the best probability
                best_probability = sum_target_per_feature
    # Print the best feature and its error
    print(f"Best feature: {best_feature}")
    print(f"Best probability: {best_probability}")

    # Retrieve the best rule for the best feature
    for value, counts in freq_table[best_feature].items():
        best_target = min(counts, key=counts.get)
        best_rule[value] = best_target 
    print(f"Best rule: {best_rule}")
    
    # Return the best feature and its rule
    return best_feature, best_rule

def test_one_r_model(data, target, best_feature, model):
    # Test the OneR model using the frequency table
    total_count = 0
    error = 0
    for row in data:
        total_count += 1
        feature_value = row[best_feature]
        predicted_class = model[feature_value]
        real_class = row[target]
        print(f"Predicted: {predicted_class}, Real: {real_class}")
        if predicted_class == real_class:
            print("Correct prediction!")
        else:
            error += 1
            print("Incorrect prediction.")
    print(f"Total predictions: {total_count}")
    print(f"Total errors: {error}")
    print(f"Total predictions: {error/total_count*100}%")

def main():
    print_dic_list(dic_list)
    train_data, test_data = split_data(dic_list, 20)
    print_dic_list(train_data)
    print_dic_list(test_data)
    freq_table = build_frequency_table(train_data, 'Class')
    feature, model = train_one_r(freq_table)
    test_one_r_model(test_data, 'Class', feature, model)

main()

