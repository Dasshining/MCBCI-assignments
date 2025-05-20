import random
import openpyxl

# Cargar el archivo y la hoja
wb = openpyxl.load_workbook("golf-dataset-categorical.xlsx")
ws = wb.active

# Obtener los encabezados (primera fila)
headers = [cell.value for cell in ws[1]]

# Crear lista de diccionarios para cada fila de datos
dic_list = []
for row in ws.iter_rows(min_row=2, values_only=True):
    dic = dict(zip(headers, row))
    dic_list.append(dic)

print(dic_list)

'''data = [
    {'Weather': 'Sunny', 'Temperature': 'Hot', 'Play': 'No'},
    {'Weather': 'Sunny', 'Temperature': 'Cool', 'Play': 'No'},
    {'Weather': 'Overcast', 'Temperature': 'Hot', 'Play': 'Yes'},
    {'Weather': 'Rain', 'Temperature': 'Cool', 'Play': 'Yes'},
    {'Weather': 'Rain', 'Temperature': 'Hot', 'Play': 'Yes'},
    {'Weather': 'Overcast', 'Temperature': 'Cool', 'Play': 'Yes'},
    {'Weather': 'Overcast', 'Temperature': 'Cool', 'Play': 'Yes'}
]'''

data = dic_list

# Definir el modelo de reglas
model = [
    {'Outlook'},  # atributo a usar
    {'Outlook': 'Sunny', 'Class': 'No'},
    {'Outlook': 'Overcast', 'Class': 'Yes'},
    {'Outlook': 'Rain', 'Class': 'No'},
]

# Seleccionar una línea aleatoria
#row = random.choice(data)
#print("Línea seleccionada:", row)

num_comparisons = 5 

for i in range(num_comparisons):
    row = random.choice(data)
    print(f"\nComparación {i+1}:")
    print("Línea seleccionada:", row)

# Buscar la regla correspondiente en el modelo
    Outlook_value = row['Outlook']
    predicted_class = None
    for rule in model[1:]:
        if rule['Outlook'] == Outlook_value:
            predicted_class = rule['Class']
            break

# Comparar el valor predicho con el real
    real_class = row['Class']
    print(f"Predicción del modelo: {predicted_class}, Valor real: {real_class}")

    if predicted_class == real_class:
        print("¡Predicción correcta!")
    else:
        print("Predicción incorrecta.")


'''def test_one_r_model(data, target, model):
    model = [
        {'Weather'},
        {'Weather': 'Sunny', 'Play': 'No'},
        {'Weather': 'Overcast', 'Play': 'Yes'},
        {'Weather': 'Rain', 'Play': 'No'},
    ]

    for row in data:
        feature_value = row[model[0]]
        model_result = row[feature_value]
        result = row[target]
        if(model_result != feature_value):
            print("Model Fail :(")'''

